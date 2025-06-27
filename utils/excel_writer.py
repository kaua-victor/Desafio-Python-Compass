import pandas as pd # Importa a biblioteca Pandas para manipulação de dados em tabelas
from openpyxl.styles import Font # Importa a classe Font para formatação de texto no Excel
from openpyxl import load_workbook # Importa o load_workbook para abrir e editar um arquivo Excel existente

# Função responsável por salvar os dados coletados no Excel
def save_excel(quotes, selic, history):

    """
    Salva os dados de cotações de moedas, SELIC e histórico do dólar em um arquivo Excel.

    Cria abas separadas para "Resumo Atual", "Taxas de Juros" e "Histórico Dólar",
    aplica formatação básica como negrito no cabeçalho e ajusta a largura das colunas.

    Args:
        quotes (dict): Dicionário com as cotações de moedas atuais.
        selic (list): Lista de dicionários com os dados da taxa SELIC.
        history (list): Lista de dicionários com o histórico do dólar.
    """

    # Cria um writer para escrever no arquivo Excel usando o openpyxl
    writer = pd.ExcelWriter('relatorio_moedas.xlsx', engine='openpyxl')

    # Aba Resumo Atual
    date_resume = []
    for currency, datas in quotes.items():
        if datas and isinstance(datas, dict):
            value = float(next(iter(datas.values()))["bid"])
            date_resume.append({"Moeda": currency, "Valor R$": round(value, 2)})
    df_resume = pd.DataFrame(date_resume)
    df_resume.to_excel(writer, sheet_name='Resumo Atual', index=False)

    # Aba Taxas de Juros
    if selic:
        df_selic = pd.DataFrame(selic)
        df_selic.columns = ['Data', 'Valor em %']
        df_selic.to_excel(writer, sheet_name='Taxas de Juros', index=False)

    # Aba Histórico Dólar
    if history:
        df_hist = pd.DataFrame(history)
        df_hist['timestamp'] = pd.to_numeric(df_hist['timestamp'], errors='coerce')
        df_hist['timestamp'] = pd.to_datetime(df_hist['timestamp'], unit='s').dt.date
        # Garante que 'bid' é numérico ANTES de arredondar
        df_hist['bid'] = pd.to_numeric(df_hist['bid'], errors='coerce')
        df_hist['bid'] = df_hist['bid'].round(2)
        df_hist = df_hist[['timestamp', 'bid']].rename(columns={'timestamp': 'Data', 'bid': 'Valor em R$'})
        df_hist.to_excel(writer, sheet_name='Histórico Dólar', index=False)

    # Fechar o writer após salvar as abas
    writer.close()  

    # Abre o arquivo para aplicar formatação de cabeçalhos e ajuste de largura de colunas
    wb = load_workbook('relatorio_moedas.xlsx')
    # Iterar por todas as planilhas
    for sheet_name in wb.sheetnames: 
        ws = wb[sheet_name]
        # Aplica negrito na primeira linha (cabeçalho)
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # Ajustar a largura da coluna 'Data' na aba 'Taxas de Juros'
        if sheet_name == 'Taxas de Juros':
            for col in ws.columns: 
                max_length = 0
                column = [cell for cell in col]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception as e:
                            print(f"Erro ao calcular largura da coluna: {e}")
                    
                # Adiciona uma margem extra para garantir que o último número não seja cortado
                adjusted_width = (max_length + 2) 
                ws.column_dimensions[col[0].column_letter].width = adjusted_width
        
        # Ajustar a largura da coluna 'Data' na aba 'Histórico Dólar'
        if sheet_name == 'Histórico Dólar':
            for col in ws.columns:
                max_length = 0
                column = [cell for cell in col]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception as e:
                        print(f"Erro ao calcular largura da coluna: {e}")

                adjusted_width = (max_length) 
                ws.column_dimensions[col[0].column_letter].width = adjusted_width
        
    # Salva após aplicar todas as alterações
    wb.save('relatorio_moedas.xlsx')